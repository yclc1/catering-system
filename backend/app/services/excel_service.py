"""Excel export service."""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from decimal import Decimal


def _create_header_style():
    return {
        "font": Font(bold=True, size=11),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "fill": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
    }


def _apply_style(cell, style_dict):
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def _cell_border():
    return Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )


def export_customer_settlement(customer_name: str, settlement, meals) -> BytesIO:
    """Generate Excel for customer monthly settlement."""
    wb = Workbook()
    ws = wb.active
    ws.title = "月度结算"

    header_style = _create_header_style()

    # Title
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"{customer_name} - {settlement.settlement_month} 月度用餐结算"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["日期", "早餐人数", "早餐金额", "午餐人数", "午餐金额", "晚餐人数", "晚餐金额", "夜宵人数", "夜宵金额", "当日合计"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        _apply_style(cell, header_style)

    # Data rows
    for row_idx, meal in enumerate(meals, 4):
        values = [
            meal.meal_date.strftime("%Y-%m-%d"),
            meal.breakfast_count, float(meal.breakfast_amount),
            meal.lunch_count, float(meal.lunch_amount),
            meal.dinner_count, float(meal.dinner_amount),
            meal.supper_count, float(meal.supper_amount),
            float(meal.daily_total),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = _cell_border()
            if isinstance(val, float):
                cell.number_format = "#,##0.00"

    # Summary row
    summary_row = len(meals) + 4
    ws.cell(row=summary_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=settlement.total_breakfast_count).font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=settlement.total_lunch_count).font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=settlement.total_dinner_count).font = Font(bold=True)
    ws.cell(row=summary_row, column=8, value=settlement.total_supper_count).font = Font(bold=True)
    ws.cell(row=summary_row, column=10, value=float(settlement.total_amount)).font = Font(bold=True)
    ws.cell(row=summary_row, column=10).number_format = "#,##0.00"

    # Adjustment and final
    ws.cell(row=summary_row + 1, column=1, value="调整金额")
    ws.cell(row=summary_row + 1, column=10, value=float(settlement.adjustment_amount)).number_format = "#,##0.00"
    ws.cell(row=summary_row + 2, column=1, value="最终结算金额").font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 2, column=10, value=float(settlement.final_amount)).font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 2, column=10).number_format = "#,##0.00"

    # Column widths
    widths = [12, 10, 12, 10, 12, 10, 12, 10, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_supplier_reconciliation(supplier_name: str, reconciliation, transactions) -> BytesIO:
    """Generate Excel for supplier monthly reconciliation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "供应商对账"

    header_style = _create_header_style()

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"{supplier_name} - {reconciliation.reconciliation_month} 月度对账单"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["单据编号", "类型", "日期", "商品", "数量", "单价", "金额"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        _apply_style(cell, header_style)

    row_idx = 4
    for txn in transactions:
        for item in txn.items:
            values = [
                txn.code,
                "入库" if txn.transaction_type == "inbound" else "退货",
                txn.transaction_date.strftime("%Y-%m-%d"),
                item.product.name if item.product else "",
                float(item.quantity),
                float(item.unit_price),
                float(item.amount),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = _cell_border()
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
            row_idx += 1

    # Summary
    ws.cell(row=row_idx + 1, column=1, value="入库合计").font = Font(bold=True)
    ws.cell(row=row_idx + 1, column=7, value=float(reconciliation.total_inbound_amount)).number_format = "#,##0.00"
    ws.cell(row=row_idx + 2, column=1, value="退货合计").font = Font(bold=True)
    ws.cell(row=row_idx + 2, column=7, value=float(reconciliation.total_return_amount)).number_format = "#,##0.00"
    ws.cell(row=row_idx + 3, column=1, value="净额").font = Font(bold=True, size=12)
    ws.cell(row=row_idx + 3, column=7, value=float(reconciliation.net_amount)).font = Font(bold=True, size=12)
    ws.cell(row=row_idx + 3, column=7).number_format = "#,##0.00"

    widths = [16, 8, 12, 20, 10, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
